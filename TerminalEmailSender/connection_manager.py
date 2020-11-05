# -*- coding: utf-8 -*-

import getpass
from smtplib import *
from message_manager import *
import socket
import time
import sys

TIME_DELAY_SEC = 0


def read_command_line_options():
    thismodule = sys.modules[__name__]

    for idx, key_val in enumerate(sys.argv, 0):
        if key_val in ['--time_delay', '-td'] and len(sys.argv) > idx + 1:
            try:
                thismodule.TIME_DELAY_SEC = int(sys.argv[idx + 1].strip())
            except ValueError as val_err:
                sys.exit("Wrong value for time delay")


class ConnectionManager:
    def __init__(self):
        self.sender_address = None
        self.psw = None
        self.sender_server = None
        self.server_port = None
        self.receiver_address = None
        self.server = None
        self.subject = None
        self.text = None

        self.set_credentials()

    def set_credentials(self):
        read_command_line_options()
        self.sender_address = input("Welcome, please input your email address (usually username@server): ")
        self.psw = getpass.getpass("Please, input the password for " + self.sender_address + ": ", stream=None)
        self.sender_server = input("Please, input the server for " + self.sender_address + ": ")

        while True:
            try:
                self.server_port = int(input(
                    "Please, input the server port for " + self.sender_address + " on " + self.sender_server + ": "))
            except ValueError:
                self.server_port = None
            else:
                break

        self.receiver_address = input("Welcome, please input the receiver's email address: ")
        self.subject = input("Welcome, please input email subject: ")
        self.text = input("Welcome, please input email corpus: ")

        # print(TIME_DELAY_SEC)

    def connect_to_server(self):
        time.sleep(TIME_DELAY_SEC)

        while True:
            try:
                self.server = SMTP(self.sender_server, self.server_port, timeout=5)
            except (SMTPException, SMTPResponseException, socket.error):
                print("An error occurred at login time: check if the sender server and the port are correct.")

            try:
                self.server.starttls()
            except SMTPException:
                print("An exception occurred with the server starting")

            try:
                self.server.login(self.sender_address, self.psw)
            except SMTPException:
                print(
                    "An error occurred at login time: check if the address " + self.sender_address +
                    " and the password are correct.")
            else:
                print("Login successful.")
                break

    def send_message(self):
        msg_man = MessageManager()
        msg_man.initialize_message(sub=self.subject, txt=self.text, from_=self.sender_address, to=self.receiver_address)
        while True:
            try:
                self.server.sendmail(self.sender_address, self.receiver_address, msg_man.compose_message())
            except SMTPException:
                print("Message not sent.")
            else:
                print("Message sent.")
                break

    def disconnect(self):
        self.server.quit()


"""#-*- coding: utf-8 -*-

import openpyxl
import smtplib
import getpass

psw = getpass.getpass("Password for mromanel@lix.polytechnique.fr: ")

server = smtplib.SMTP('smtp.lix.polytechnique.fr', 465)
server.starttls()
server.login("mromanel@lix.polytechnique.fr", psw)

bookGrades = openpyxl.load_workbook('/Users/marcoromanelli/Desktop/PhD_documents/grades/Midterm_grades_harmonises.xlsx',data_only=True)
bookMails = openpyxl.load_workbook('/Users/marcoromanelli/Desktop/PhD_documents/grades/CSE103_notes_partiel_initial.xlsx',data_only=True)

sheetGrades = bookGrades.active

sheetMails = bookMails.active

senderAddress="mromanel@lix.polytechnique.fr"

for i in range(2, 72):
    print i
    if i == 2 or i == 71:
        print sheetMails.cell(row=i, column=1).value

    if str(sheetGrades.cell(row=i, column=5).value) == "absent":
        continue

    else:
        Surname = sheetMails.cell(row=i, column=1).value.encode('utf-8')
        Surname_= sheetGrades.cell(row=i, column=1).value.encode('utf-8')
        #print(Surname)
        #print(Surname_)
        Name = sheetMails.cell(row=i, column=2).value.encode('utf-8')
        Name_ = sheetGrades.cell(row=i, column=2).value.encode('utf-8')
        #print(Name)
        #print(Name_)
        Matr = sheetMails.cell(row=i, column=4).value.encode('utf-8')
        Matr_ = sheetGrades.cell(row=i, column=3).value.encode('utf-8')
        #print(Matr)
        #print(Matr_)

        if Surname != Surname_ or Name != Name_ or Matr != Matr_:
            print "ERROR"
            exit()

        #destAddress = sheetMails.cell(row=i, column=3).value

        destAddress = "marco.romanelli@inria.fr"

        harmonizedGrade = round(sheetGrades.cell(row=i, column=14).value, 2)

        Q11 = round(sheetGrades.cell(row=i, column=6).value, 2)

        Q12 = round(sheetGrades.cell(row=i, column=7).value, 2)

        Q13 = round(sheetGrades.cell(row=i, column=8).value, 2)

        Q14 = round(sheetGrades.cell(row=i, column=9).value, 2)

        Q21 = round(sheetGrades.cell(row=i, column=10).value, 2)

        Q22 = round(sheetGrades.cell(row=i, column=11).value, 2)

        Q31 = round(sheetGrades.cell(row=i, column=12).value, 2)

        Q32 = round(sheetGrades.cell(row=i, column=13).value, 2)

        SUBJECT = "Midterm grades CSE103"
        TEXT = "Dear " + str(Name) + " " + str(Surname) + ",\n" \
               + " here is your grade after the harmonization: " + str(harmonizedGrade) + ".\n" \
               + "Please find below the score for each question:\n"\
               + "Q1.1: " + str(Q11) + "\n" \
               + "Q1.2: " + str(Q12) + "\n" \
               + "Q1.3: " + str(Q13) + "\n" \
               + "Q1.4: " + str(Q14) + "\n" \
               + "Q2.1: " + str(Q21) + "\n" \
               + "Q2.2: " + str(Q22) + "\n" \
               + "Q3.1: " + str(Q31) + "\n" \
               + "Q3.2: " + str(Q32) + "\n\n" \
               + "Best regards."

        message = 'Subject: {}\n\n{}'.format(SUBJECT, TEXT)
        server.sendmail(senderAddress, destAddress, message)
        print destAddress
server.quit()
"""
